"""Bulk SMS campaign routes — TopAI-owned audience, attestation, and jobs."""

from __future__ import annotations

import csv
import hashlib
import io
import json
import logging
import os
import re
import uuid

from flask import Blueprint, flash, jsonify, redirect, render_template, request, url_for

import auth
import config
import db
import tenant_sms_db as tdb
from sms_authorization import (
    CAMPAIGN_CERT_TEXT,
    ONE_TO_ONE_CERT_TEXT,
    check_telnyx_toll_free_send_allowed,
    get_telnyx_toll_free_verification_status,
    is_sms_sending_enabled,
    is_telnyx_toll_free_verified,
    require_tenant_sender,
)
from sms_validation import validate_e164_phone

logger = logging.getLogger(__name__)

sms_campaigns_bp = Blueprint("sms_campaigns", __name__)

BULK_VERIFICATION_BANNER = (
    "Telnyx SMS is configured. Bulk sending will become available after "
    "toll-free verification is approved."
)
WORKER_UNAVAILABLE_BANNER = "Campaign processing worker is currently unavailable."


def _auth_gate():
    """Return (user, response). response is set when the request should not proceed."""
    user = auth.get_current_user()
    if not user:
        nxt = request.path
        if request.query_string:
            nxt = f"{request.path}?{request.query_string.decode()}"
        return None, redirect(url_for("login", next=nxt))
    if not auth.user_has_active_subscription(user):
        return None, (
            render_template(
                "error.html",
                message=(
                    "An active TopAI subscription is required to use Bulk SMS. "
                    "Subscribe or renew your plan, then return here."
                ),
            ),
            402,
        )
    return user, None


def _nav(user, active_nav="sms-campaigns"):
    return {
        "email": user["email"],
        "has_billing_portal": bool(user.get("stripe_customer_id")),
        "active_nav": active_nav,
        "product_name": "TopAI Real Estate Tools",
    }


def _bulk_status_context(user_id=None):
    """Non-secret provider / verification / worker status for Bulk SMS pages.

    Must never raise: missing worker heartbeat, sender, or terms rows are
    expected empty states and must still allow the page to render.
    """
    from sms_authorization import (
        TELNYX_CONFIGURED_PENDING_MSG,
        telnyx_configuration_complete,
    )

    provider = (config.SMS_PROVIDER or "telnyx").lower().strip()
    verification = get_telnyx_toll_free_verification_status() or "unknown"
    toll_ok, _toll_err = check_telnyx_toll_free_send_allowed()
    try:
        worker_health = tdb.get_campaign_worker_health(stale_seconds=120) or {}
    except Exception:
        logger.exception("campaign worker health lookup failed")
        worker_health = {
            "state": "unknown",
            "message": WORKER_UNAVAILABLE_BANNER,
        }
    worker_ok = worker_health.get("state") == "running"
    terms_ok = True
    if user_id:
        try:
            terms_ok = tdb.has_accepted_sms_terms(user_id)
        except Exception:
            logger.exception("sms terms lookup failed user_id=%s", user_id)
            terms_ok = False
    try:
        telnyx_ready = provider != "telnyx" or telnyx_configuration_complete()
    except Exception:
        logger.exception("telnyx configuration check failed")
        telnyx_ready = False
    sending_enabled = bool(
        is_sms_sending_enabled() and toll_ok and worker_ok and terms_ok and telnyx_ready
    )
    verification_message = None
    if provider == "telnyx" and telnyx_ready and not toll_ok:
        verification_message = TELNYX_CONFIGURED_PENDING_MSG
    elif provider == "telnyx" and not toll_ok:
        verification_message = BULK_VERIFICATION_BANNER
    return {
        "sms_provider": provider,
        "telnyx_configured": telnyx_ready if provider == "telnyx" else None,
        "toll_free_number_display": getattr(config, "SMS_SUPPORT_DISPLAY", None)
        or "(888) 821-0810",
        "toll_free_verification_status": verification if provider == "telnyx" else None,
        "bulk_sending_enabled": sending_enabled,
        "toll_free_verified": is_telnyx_toll_free_verified() if provider == "telnyx" else True,
        "verification_block_message": verification_message,
        "worker_available": worker_ok,
        "worker_health_state": worker_health.get("state"),
        "worker_block_message": worker_health.get("message") if not worker_ok else None,
        "launch_controls_enabled": sending_enabled,
        "terms_ok": terms_ok,
    }


def _safe_require_sender(user_id):
    try:
        return require_tenant_sender(user_id)
    except Exception:
        logger.exception("tenant sender lookup failed user_id=%s", user_id)
        return None, "SMS sender status is temporarily unavailable."


def _render_merge(template, fields, defaults=None):
    defaults = defaults or {}
    text = template or ""

    def repl(match):
        key = match.group(1).strip().lower()
        val = fields.get(key) or defaults.get(key) or ""
        return str(val)

    return re.sub(r"\[([a-zA-Z0-9_]+)\]", repl, text)


@sms_campaigns_bp.route("/crm/sms-campaigns")
def campaigns_list():
    user, gate = _auth_gate()
    if gate:
        return gate
    try:
        campaigns = tdb.list_campaigns(user["id"]) or []
    except Exception:
        logger.exception("list_campaigns failed user_id=%s", user["id"])
        campaigns = []
        flash("Campaign list is temporarily unavailable. Please try again.")
    sender, sender_err = _safe_require_sender(user["id"])
    status = _bulk_status_context(user["id"])
    return render_template(
        "crm_sms_campaigns.html",
        campaigns=campaigns,
        sender=sender,
        sender_err=sender_err,
        **status,
        **_nav(user),
    )


@sms_campaigns_bp.route("/crm/sms-campaigns/new", methods=["GET", "POST"])
def campaign_new():
    user, gate = _auth_gate()
    if gate:
        return gate
    sender, sender_err = _safe_require_sender(user["id"])
    status = _bulk_status_context(user["id"])
    form_title = (request.form.get("title") or "").strip()[:200]
    form_purpose = (request.form.get("campaign_purpose") or "real_estate_follow_up").strip()
    error = None

    if request.method == "POST":
        title = form_title or "Untitled campaign"
        purpose = form_purpose or "real_estate_follow_up"
        try:
            # Draft create is allowed while send/launch is blocked (e.g. pending
            # toll-free verification). Launch remains gated on the detail page.
            cid = tdb.create_campaign(user["id"], title, campaign_purpose=purpose)
            tdb.append_sms_audit(
                user["id"], "campaign_created", actor_user_id=user["id"], campaign_id=cid
            )
            return redirect(url_for("sms_campaigns.campaign_detail", campaign_id=cid))
        except Exception:
            logger.exception("create_campaign failed user_id=%s", user["id"])
            error = (
                "Could not create the campaign draft. Please try again. "
                "Sending remains blocked until compliance checks pass."
            )

    return render_template(
        "crm_sms_campaign_new.html",
        error=error,
        form_title=form_title,
        form_purpose=form_purpose,
        sender=sender,
        sender_err=sender_err,
        **status,
        **_nav(user),
    )


@sms_campaigns_bp.route("/crm/sms-campaigns/<int:campaign_id>", methods=["GET", "POST"])
def campaign_detail(campaign_id):
    user, gate = _auth_gate()
    if gate:
        return gate
    campaign = tdb.get_campaign(campaign_id, user["id"])
    if not campaign:
        return redirect(url_for("sms_campaigns.campaigns_list"))
    error = None
    sender, sender_err = _safe_require_sender(user["id"])
    bulk_status = _bulk_status_context(user["id"])

    if request.method == "POST":
        action = (request.form.get("action") or "").strip()
        if action == "save_message":
            template = (request.form.get("message_template") or "").strip()[:1000]
            purpose = (request.form.get("campaign_purpose") or campaign.get("campaign_purpose") or "").strip()
            old_fp = campaign.get("content_fingerprint")
            new_fp = hashlib.sha256(f"{template}|{purpose}".encode()).hexdigest()
            tdb.update_campaign(
                campaign_id,
                user["id"],
                message_template=template,
                campaign_purpose=purpose,
                content_fingerprint=new_fp,
                sender_number=(sender or {}).get("sender_number"),
            )
            if old_fp and old_fp != new_fp:
                tdb.invalidate_campaign_attestations(user["id"], campaign_id)
                flash("Message changed — recertification required before launch.")
            tdb.append_sms_audit(user["id"], "campaign_edited", actor_user_id=user["id"], campaign_id=campaign_id)
            return redirect(url_for("sms_campaigns.campaign_detail", campaign_id=campaign_id))

        if action == "import_audience":
            error = _import_audience(user["id"], campaign_id, request)
            if not error:
                flash("Audience imported. Review exclusions, then certify before launch.")
                return redirect(url_for("sms_campaigns.campaign_detail", campaign_id=campaign_id))

        if action == "certify":
            if not request.form.get("attestation_accepted"):
                error = "You must check the campaign certification checkbox."
            else:
                campaign = tdb.get_campaign(campaign_id, user["id"])
                recipients = tdb.list_campaign_recipients(campaign_id, user["id"])
                eligible = [r for r in recipients if r.get("eligible")]
                excluded = len(recipients) - len(eligible)
                if not campaign.get("message_template"):
                    error = "Compose a message before certifying."
                elif not eligible:
                    error = "No eligible recipients to certify."
                elif sender_err:
                    error = sender_err
                else:
                    att_id = tdb.create_campaign_attestation(
                        user["id"],
                        user["id"],
                        campaign_id,
                        eligible_count=len(eligible),
                        excluded_count=excluded,
                        campaign_purpose=campaign.get("campaign_purpose") or "campaign",
                        message_body=campaign.get("message_template") or "",
                        audience_snapshot_id=campaign.get("audience_snapshot_id") or "",
                        provider=config.SMS_PROVIDER,
                        scheduled_launch_at=request.form.get("scheduled_at") or None,
                    )
                    tdb.update_campaign(campaign_id, user["id"], attestation_id=att_id)
                    tdb.append_sms_audit(
                        user["id"],
                        "consent_certification_accepted",
                        actor_user_id=user["id"],
                        campaign_id=campaign_id,
                        metadata={"attestation_id": att_id, "eligible": len(eligible)},
                    )
                    flash("Audience certified (subscriber certification — not TopAI verified).")
                    return redirect(url_for("sms_campaigns.campaign_detail", campaign_id=campaign_id))

        if action == "launch":
            campaign = tdb.get_campaign(campaign_id, user["id"])
            toll_ok, toll_err = check_telnyx_toll_free_send_allowed()
            if not toll_ok:
                error = toll_err or BULK_VERIFICATION_BANNER
            elif not bulk_status["worker_available"]:
                error = WORKER_UNAVAILABLE_BANNER
            elif config.TELNYX_TRIAL_MODE and (config.SMS_PROVIDER or "").lower() == "telnyx":
                recipients = tdb.list_campaign_recipients(campaign_id, user["id"], eligible_only=True)
                verified = "".join(c for c in (config.TELNYX_VERIFIED_TEST_NUMBER or "") if c.isdigit())
                bad = [
                    r
                    for r in recipients
                    if "".join(c for c in (r.get("phone_number") or "") if c.isdigit()) != verified
                ]
                if bad or not recipients:
                    error = (
                        "Telnyx trial mode: campaign audience must contain only the verified test phone number."
                    )
                else:
                    error = None
            else:
                error = None
            if not error:
                att = tdb.get_valid_campaign_attestation(
                    user["id"],
                    campaign_id,
                    message_body=campaign.get("message_template") or "",
                    audience_snapshot_id=campaign.get("audience_snapshot_id") or "",
                    purpose=campaign.get("campaign_purpose") or "campaign",
                )
                if not att:
                    error = "Certify the final audience and message before launch."
                elif sender_err:
                    error = sender_err
                else:
                    scheduled = (request.form.get("scheduled_at") or "").strip() or None
                    tdb.create_jobs_for_campaign(campaign_id, user["id"])
                    if scheduled:
                        tdb.update_campaign(
                            campaign_id, user["id"], status="scheduled", scheduled_at=scheduled
                        )
                        tdb.append_sms_audit(
                            user["id"], "campaign_scheduled", actor_user_id=user["id"], campaign_id=campaign_id
                        )
                    else:
                        from datetime import datetime, timezone

                        tdb.update_campaign(
                            campaign_id,
                            user["id"],
                            status="processing",
                            started_at=datetime.now(timezone.utc).isoformat(),
                        )
                        tdb.append_sms_audit(
                            user["id"], "campaign_launched", actor_user_id=user["id"], campaign_id=campaign_id
                        )
                    flash("Campaign queued. The worker sends messages outside this request.")
                    return redirect(url_for("sms_campaigns.campaign_detail", campaign_id=campaign_id))

        if action in {"pause", "resume", "cancel"}:
            mapping = {
                "pause": ("paused", "campaign_paused"),
                "resume": ("processing", "campaign_resumed"),
                "cancel": ("cancelled", "campaign_cancelled"),
            }
            new_status, audit = mapping[action]
            tdb.update_campaign(campaign_id, user["id"], status=new_status)
            if action == "cancel":
                with db.get_db() as conn:
                    conn.execute(
                        """
                        UPDATE sms_campaign_jobs SET status='cancelled', updated_at=?
                        WHERE campaign_id=? AND user_id=? AND status IN ('pending','claimed')
                        """,
                        (
                            __import__("datetime").datetime.now(
                                __import__("datetime").timezone.utc
                            ).isoformat(),
                            campaign_id,
                            user["id"],
                        ),
                    )
            tdb.append_sms_audit(user["id"], audit, actor_user_id=user["id"], campaign_id=campaign_id)
            return redirect(url_for("sms_campaigns.campaign_detail", campaign_id=campaign_id))

    campaign = tdb.get_campaign(campaign_id, user["id"])
    recipients = tdb.list_campaign_recipients(campaign_id, user["id"])
    job_stats = tdb.count_jobs_by_status(campaign_id, user["id"])
    return render_template(
        "crm_sms_campaign_detail.html",
        campaign=campaign,
        recipients=recipients[:200],
        recipient_count=len(recipients),
        eligible_count=sum(1 for r in recipients if r.get("eligible")),
        job_stats=job_stats,
        error=error,
        sender=sender,
        sender_err=sender_err,
        campaign_cert_text=CAMPAIGN_CERT_TEXT,
        warning_import=(
            "Upload only contacts who have authorized you or your business to text them. "
            "Importing or possessing a phone number does not establish consent."
        ),
        **bulk_status,
        **_nav(user),
    )


def _import_audience(user_id, campaign_id, request):
    upload = request.files.get("audience_file")
    if not upload or not upload.filename:
        return "Upload a CSV or Excel file."
    name = upload.filename.lower()
    data = upload.read(config.SMS_IMPORT_MAX_BYTES + 1)
    if len(data) > config.SMS_IMPORT_MAX_BYTES:
        return "File exceeds size limit."
    rows = []
    try:
        if name.endswith(".csv"):
            text = data.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            rows = list(reader)
        elif name.endswith(".xlsx"):
            import openpyxl

            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip() for c in next(ws.iter_rows(max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append({headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))})
        else:
            return "Supported formats: .csv, .xlsx"
    except Exception as exc:
        logger.warning("Audience import parse failed: %s", type(exc).__name__)
        return "Could not parse the uploaded file."

    if len(rows) > config.SMS_IMPORT_MAX_ROWS:
        return f"Import limited to {config.SMS_IMPORT_MAX_ROWS} rows."

    recipients = []
    seen = set()
    for row in rows:
        phone_raw = (
            row.get("phone")
            or row.get("phone_number")
            or row.get("mobile")
            or row.get("Phone")
            or ""
        )
        phone, err = validate_e164_phone(str(phone_raw))
        first = str(row.get("first_name") or row.get("firstName") or "").strip()
        last = str(row.get("last_name") or row.get("lastName") or "").strip()
        full = str(row.get("name") or row.get("full_name") or f"{first} {last}").strip()
        merge = {
            "first_name": first,
            "last_name": last,
            "full_name": full,
            "email": str(row.get("email") or "").strip(),
            "property_interest": str(row.get("property_interest") or "").strip(),
        }
        exclusion = None
        eligible = True
        lead_id = None
        if err or not phone:
            exclusion = "invalid_number"
            eligible = False
            phone = str(phone_raw)[:32]
        elif phone in seen:
            exclusion = "duplicate"
            eligible = False
        else:
            seen.add(phone)
            if tdb.is_suppressed(user_id, phone):
                exclusion = "suppressed"
                eligible = False
            lead = db.get_lead_by_phone(user_id, phone)
            if lead:
                lead_id = lead["id"]
                if (lead.get("opt_out_status") or "") == "opted_out" or (
                    lead.get("sms_consent_status") or ""
                ) == "opted_out":
                    exclusion = "opted_out"
                    eligible = False
            else:
                # Create lead as not_certified
                from external_leads.ingest import ingest_external_lead

                result = ingest_external_lead(
                    user_id,
                    {
                        "full_name": full or "Imported Contact",
                        "phone": phone,
                        "email": merge.get("email"),
                        "property_interest": merge.get("property_interest"),
                    },
                    method="csv",
                    actor_user_id=user_id,
                )
                lead_id = result.get("lead_id")
                if lead_id:
                    import external_leads_db as xdb

                    xdb.set_lead_sms_consent_state(
                        lead_id,
                        user_id,
                        sms_consent_status="not_certified",
                        sms_sending_blocked=True,
                        actor_user_id=user_id,
                        source="campaign_import",
                    )
        recipients.append(
            {
                "lead_id": lead_id,
                "phone_number": phone,
                "merge_fields": merge,
                "eligible": eligible,
                "exclusion_reason": exclusion,
            }
        )

    tdb.replace_campaign_recipients(campaign_id, user_id, recipients)
    tdb.append_sms_audit(
        user_id,
        "import_completed",
        actor_user_id=user_id,
        campaign_id=campaign_id,
        metadata={"rows": len(recipients)},
    )
    return None


@sms_campaigns_bp.route("/crm/sms-diagnostics", methods=["GET", "POST"])
def sms_diagnostics():
    user, gate = _auth_gate()
    if gate:
        return gate
    if request.method == "POST" and request.form.get("action") == "accept_terms":
        tdb.accept_sms_terms(
            user["id"],
            user["id"],
            ip_address=request.headers.get("X-Forwarded-For", request.remote_addr),
            user_agent=request.headers.get("User-Agent"),
        )
        flash("SMS terms accepted.")
        return redirect(url_for("sms_campaigns.sms_diagnostics"))

    from sms_providers import get_sms_provider

    provider = get_sms_provider()
    sender, sender_err = _safe_require_sender(user["id"])
    last_out = tdb.latest_sms_event(user["id"], direction="outbound")
    last_in = tdb.latest_sms_event(user["id"], direction="inbound")
    bulk = _bulk_status_context(user["id"])
    info = {
        "active_provider": config.SMS_PROVIDER,
        "provider_configured": provider.is_configured(),
        "sender": sender,
        "sender_error": sender_err,
        "token_configured": bool(
            config.TELNYX_API_KEY
            if (config.SMS_PROVIDER or "").lower() == "telnyx"
            else config.SIMPLETEXTING_API_TOKEN
        ),
        "webhook_secret_configured": bool(
            config.TELNYX_PUBLIC_KEY
            if (config.SMS_PROVIDER or "").lower() == "telnyx"
            else config.SIMPLETEXTING_WEBHOOK_SECRET
        ),
        "telnyx_trial_mode": bool(config.TELNYX_TRIAL_MODE)
        if (config.SMS_PROVIDER or "").lower() == "telnyx"
        else False,
        "telnyx_profile_configured": bool(config.TELNYX_MESSAGING_PROFILE_ID),
        "telnyx_phone_configured": bool(config.TELNYX_PHONE_NUMBER),
        "telnyx_verified_test_configured": bool(config.TELNYX_VERIFIED_TEST_NUMBER),
        "telnyx_public_key_configured": bool(config.TELNYX_PUBLIC_KEY),
        "toll_free_verification_status": bulk.get("toll_free_verification_status"),
        "bulk_sending_enabled": bulk.get("bulk_sending_enabled"),
        "worker_available": bulk.get("worker_available"),
        "worker_health_state": bulk.get("worker_health_state"),
        "worker_block_message": bulk.get("worker_block_message"),
        "verification_block_message": bulk.get("verification_block_message"),
        "toll_free_number_display": bulk.get("toll_free_number_display"),
        "telnyx_configured": bulk.get("telnyx_configured"),
        "webhook_route": "/webhooks/telnyx/messaging",
        "last_outbound": last_out,
        "last_inbound": last_in,
        "queue_backlog": tdb.count_pending_campaign_jobs(),
        "terms_ok": bulk.get("terms_ok"),
        "terms_version": config.SMS_TERMS_VERSION,
        "one_to_one_cert_text": ONE_TO_ONE_CERT_TEXT,
        "campaign_cert_text": CAMPAIGN_CERT_TEXT,
        "sender_info": provider.get_sender_information(),
        "warning": (
            "You may send SMS only to contacts who have authorized you or your business "
            "to text them. TopAI does not independently verify your supporting consent records."
        ),
    }
    return render_template("crm_sms_diagnostics.html", **info, **_nav(user, "sms-diagnostics"))


@sms_campaigns_bp.route("/crm/sms-campaigns/<int:campaign_id>/export")
def campaign_export(campaign_id):
    user, gate = _auth_gate()
    if gate:
        return gate
    campaign = tdb.get_campaign(campaign_id, user["id"])
    if not campaign:
        return redirect(url_for("sms_campaigns.campaigns_list"))
    recipients = tdb.list_campaign_recipients(campaign_id, user["id"])
    jobs = {}
    with db.get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM sms_campaign_jobs WHERE campaign_id = ? AND user_id = ?",
            (campaign_id, user["id"]),
        ).fetchall()
        for r in rows:
            jobs[r["recipient_id"]] = dict(r)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(
        [
            "phone_number",
            "eligible",
            "exclusion_reason",
            "job_status",
            "provider_message_id",
            "failure_message",
        ]
    )
    for r in recipients:
        job = jobs.get(r["id"]) or {}
        writer.writerow(
            [
                r.get("phone_number"),
                r.get("eligible"),
                r.get("exclusion_reason") or "",
                job.get("status") or "",
                job.get("provider_message_id") or "",
                job.get("failure_message") or "",
            ]
        )
    from flask import Response

    return Response(
        buf.getvalue(),
        mimetype="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=campaign_{campaign_id}_export.csv"
        },
    )


@sms_campaigns_bp.route("/r/<token>")
def tracking_redirect(token):
    row = tdb.record_link_click(token)
    if not row:
        return "Not found", 404
    return redirect(row["destination_url"])
